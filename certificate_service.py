import os
import io
import zipfile
import csv
import tempfile
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

# ── Font cache: (path, size) -> ImageFont object ─────────────────────────────
_font_cache = {}

def _get_font(font_path, size):
    key = (font_path, size)
    if key not in _font_cache:
        _font_cache[key] = ImageFont.truetype(font_path, size) if font_path else ImageFont.load_default()
    return _font_cache[key]

# ── Font path cache: name -> path string ─────────────────────────────────────
_font_path_cache = {}

def _get_font_path(font_name):
    if font_name not in _font_path_cache:
        path = None
        for candidate in [f"fonts/{font_name}.ttf", "default_font.ttf", "font.ttf"]:
            if os.path.exists(candidate):
                path = candidate
                break
        _font_path_cache[font_name] = path
    return _font_path_cache[font_name]

# ── Font-fit cache: (font_path, max_size, min_size, max_width, text_len) -> size
# Names of the same length almost always fit the same font size, so we skip
# the binary search entirely for repeated lengths — massive win on free tier.
_fit_cache = {}

def _fit_font_size(draw, text, font_path, max_font_size, min_font_size, max_text_width):
    if not font_path:
        return ImageFont.load_default()

    cache_key = (font_path, max_font_size, min_font_size, max_text_width, len(text))
    if cache_key in _fit_cache:
        return _get_font(font_path, _fit_cache[cache_key])

    lo, hi = min_font_size, max_font_size
    best = min_font_size
    while lo <= hi:
        mid = (lo + hi) // 2
        font = _get_font(font_path, mid)
        bbox = draw.textbbox((0, 0), text, font=font)
        if (bbox[2] - bbox[0]) <= max_text_width:
            best = mid; lo = mid + 1
        else:
            hi = mid - 1

    _fit_cache[cache_key] = best
    return _get_font(font_path, best)

# ── Helpers ───────────────────────────────────────────────────────────────────
def extract_data_rows(data_bytes, filename):
    ext = os.path.splitext(filename)[1].lower()
    headers, data_rows = [], []

    if ext == '.txt':
        lines = [l.strip() for l in data_bytes.decode('utf-8').split('\n') if l.strip()]
        return ["Line_1"], [{"Line_1": l} for l in lines]

    elif ext == '.csv':
        reader = csv.reader(io.StringIO(data_bytes.decode('utf-8')))
        rows = list(reader)
        if rows:
            headers = [str(h).strip() or f"Column_{i+1}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                padded = row + [''] * max(0, len(headers) - len(row))
                d = {headers[i]: padded[i].strip() for i in range(len(headers))}
                if any(d.values()):
                    data_rows.append(d)

    elif ext in ('.xlsx', '.xls'):
        wb = load_workbook(io.BytesIO(data_bytes), data_only=True, read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        if rows:
            headers = [str(h).strip() if h is not None else f"Column_{i+1}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                padded = list(row) + [None] * max(0, len(headers) - len(row))
                d = {headers[i]: (str(padded[i]).strip() if padded[i] is not None else "") for i in range(len(headers))}
                if any(d.values()):
                    data_rows.append(d)

    return headers, data_rows


def hex_to_rgb(hex_str):
    h = hex_str.lstrip('#')
    try:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        return (255, 255, 255)


def draw_elements_on_image(image, prebuilt_elements, row_data):
    """prebuilt_elements is a list of dicts with resolved font_path and rgb colors."""
    draw = ImageDraw.Draw(image)

    for el in prebuilt_elements:
        text = row_data.get(el['column'], '') or el.get('literal', '')
        if not text:
            continue
        text = str(text).upper()

        font = _fit_font_size(draw, text, el['font_path'],
                              el['max_font_size'], el['min_font_size'], el['max_text_width'])

        bbox = draw.textbbox((0, 0), text, font=font)
        text_w = bbox[2] - bbox[0]

        align = el.get('align', 'left')
        if align == 'center':
            draw_x = el['text_x'] - text_w // 2
        elif align == 'right':
            draw_x = el['text_x'] - text_w
        else:
            draw_x = el['text_x']

        draw_y = el['text_y']

        if el.get('shadow'):
            draw.text((draw_x + 2, draw_y + 2), text, fill=(0, 0, 0), font=font)

        sw = el.get('stroke_width', 0)
        if sw > 0:
            sc = el['stroke_rgb']
            for dx in range(-sw, sw + 1):
                for dy in range(-sw, sw + 1):
                    if dx or dy:
                        draw.text((draw_x + dx, draw_y + dy), text, fill=sc, font=font)

        draw.text((draw_x, draw_y), text, fill=el['text_rgb'], font=font)


def _prebuilt(elements):
    """Resolve font paths and RGB colors once before the loop — not per row."""
    result = []
    for el in elements:
        font_name = el.get('font', 'Roboto')
        result.append({
            'column':        el.get('column', ''),
            'literal':       el.get('text', ''),
            'font_path':     _get_font_path(font_name),
            'text_rgb':      hex_to_rgb(el.get('text_color', '#ffffff')),
            'stroke_rgb':    hex_to_rgb(el.get('stroke_color', '#000000')),
            'text_x':        int(el.get('text_x', 100)),
            'text_y':        int(el.get('text_y', 390)),
            'max_font_size': int(el.get('max_font_size', 43)),
            'min_font_size': int(el.get('min_font_size', 30)),
            'max_text_width':int(el.get('max_text_width', 900)),
            'align':         el.get('align', 'left'),
            'shadow':        bool(el.get('shadow', False)),
            'stroke_width':  int(el.get('stroke_width', 0)),
        })
    return result


def _resize_template(template_bytes, max_side=1500):
    """Downscale large templates. Smaller = faster per-cert processing on free tier."""
    img = Image.open(io.BytesIO(template_bytes)).convert("RGB")
    w, h = img.size
    if max(w, h) > max_side:
        ratio = max_side / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    return img


def _primary_col(elements, row):
    col = elements[0].get('column') if elements else None
    if col and col in row:
        return col
    for common in ("Name", "Participant", "Full Name", "Participant Name"):
        if common in row:
            return common
    return list(row.keys())[0]


def _safe_name(val):
    return "".join(c for c in str(val).upper() if c.isalnum() or c in " _-").strip()


# ── Per-certificate render (runs in thread) ───────────────────────────────────
def _render_one(args):
    row, template_image, pbuilt = args
    img = template_image.copy()
    draw_elements_on_image(img, pbuilt, row)
    buf = io.BytesIO()
    # subsampling=2 (4:2:0) is ~20% faster than subsampling=0 on CPU-limited hosts.
    # optimize=False avoids a second encoding pass — no visual difference, saves time.
    img.save(buf, format='JPEG', quality=82, subsampling=2, optimize=False)
    return buf.getvalue()


# ── Main generation ───────────────────────────────────────────────────────────
def generate_certificates_zip_to_file(template_bytes, data_bytes, data_filename, config):
    template_image = _resize_template(template_bytes)
    headers, data_rows = extract_data_rows(data_bytes, data_filename)

    raw_elements = config.get('elements', [])
    pbuilt = _prebuilt(raw_elements)   # resolve paths/colors once

    # Pre-warm font cache for all elements at max size so first row isn't slow
    dummy_draw = ImageDraw.Draw(template_image.copy())
    for el in pbuilt:
        if el['font_path']:
            _get_font(el['font_path'], el['max_font_size'])

    # Build filenames up front
    seen_names = set()
    filenames = []
    for i, row in enumerate(data_rows):
        col = _primary_col(raw_elements, row)
        sn = _safe_name(row.get(col, f"cert_{i+1}")) or f"cert_{i+1}"
        base, n = sn, 1
        while sn in seen_names:
            sn = f"{base}_{n}"; n += 1
        seen_names.add(sn)
        filenames.append(sn)

    # Render all images in parallel.
    # 2 threads is optimal for Render free tier (0.1 CPU).
    # PIL releases the GIL during JPEG encoding so threads genuinely overlap.
    args_list = [(row, template_image, pbuilt) for row in data_rows]
    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(_render_one, args_list))

    # Write ZIP to disk (avoids RAM spikes)
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    with zipfile.ZipFile(temp_zip, "w", zipfile.ZIP_STORED) as zf:
        for i, img_bytes in enumerate(results):
            zf.writestr(f"{filenames[i]}.jpg", img_bytes)

    temp_zip.close()
    return temp_zip.name, len(data_rows), os.path.getsize(temp_zip.name)


# ── Preview ───────────────────────────────────────────────────────────────────
def generate_single_preview(template_bytes, config, preview_row):
    template_image = _resize_template(template_bytes)
    pbuilt = _prebuilt(config.get('elements', []))

    for el in pbuilt:
        if el['column'] and not preview_row.get(el['column']):
            preview_row[el['column']] = "MULEARNVJC"

    img = template_image.copy()
    draw_elements_on_image(img, pbuilt, preview_row)

    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=82, subsampling=2, optimize=False)
    return buf.getvalue()
