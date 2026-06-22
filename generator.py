from PIL import Image, ImageDraw, ImageFont
from PIL.PngImagePlugin import PngInfo
from openpyxl import load_workbook
import os

# =======================
# CONFIG
# =======================
TEMPLATE_IMAGE = "certificate.png"   # Certificate template (PNG)
EXCEL_FILE = "names.xlsx"            # Excel file
FONT_FILE = "font.ttf"               # Font file
OUTPUT_DIR = "certificates"

MAX_FONT_SIZE = 43
MIN_FONT_SIZE = 30
MAX_TEXT_WIDTH = 900                 # Max width for name text (px)

TEXT_COLOR = (255, 255, 255)         # White
TEXT_X = int(100)
TEXT_Y = int(390)

# =======================
# SETUP
# =======================
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = load_workbook(EXCEL_FILE)
sheet = wb.active

# =======================
# GENERATE CERTIFICATES
# =======================
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue

    # ALL CAPS name
    raw_name = str(row[0]).strip()
    name = " ".join(raw_name.split()).upper()

    # Load image
    image = Image.open(TEMPLATE_IMAGE).convert("RGB")
    draw = ImageDraw.Draw(image)

    x = int(round(TEXT_X))
    y = int(round(TEXT_Y))

    # -----------------------
    # AUTO FONT RESIZE
    # -----------------------
    font_size = MAX_FONT_SIZE
    font = ImageFont.truetype(FONT_FILE, font_size)

    bbox = draw.textbbox((0, 0), name, font=font)
    text_width = bbox[2] - bbox[0]

    while text_width > MAX_TEXT_WIDTH and font_size > MIN_FONT_SIZE:
        font_size -= 2
        font = ImageFont.truetype(FONT_FILE, font_size)
        bbox = draw.textbbox((0, 0), name, font=font)
        text_width = bbox[2] - bbox[0]

    # Draw name
    draw.text(
        (x, y),
        name,
        fill=TEXT_COLOR,
        font=font
    )

    # -----------------------
    # METADATA
    # -----------------------
    meta = PngInfo()
    meta.add_text("Title", "Certificate of Participation")
    meta.add_text("Participant Name", name)
    meta.add_text("Issuer", "μLearn VJCET")

    # Safe filename
    safe_name = "".join(c for c in name if c.isalnum() or c in " _-")
    output_path = os.path.join(OUTPUT_DIR, f"{safe_name}.png")

    image.save(output_path, pnginfo=meta)
    print(f"✔ Generated: {safe_name}.png")

print("\n🎉 All certificates generated successfully!")
